"""
Scraper assistido Google Flights — SSA -> NAT (somente ida)
Projeto: coleta de população + sorteio de amostra

Requisitos:
    pip install selenium openpyxl webdriver-manager

Como usar:
    python scraper_ssa_nat_8datas_150dias_id_voo.py

Observação importante:
    Este script faz coleta assistida. Se a automação não conseguir trocar a data
    no Google Flights, ele pausa e pede para você ajustar a data manualmente.
    Isso evita perder uma data da pesquisa por causa de mudança na interface.
"""

import random
import re
import time
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from selenium import webdriver
from selenium.common.exceptions import ElementClickInterceptedException, StaleElementReferenceException, TimeoutException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager

# =========================
# CONFIGURAÇÕES DO ESTUDO
# =========================

ORIGEM = "SSA"
DESTINO = "NAT"
ARQUIVO_SAIDA = "populacao_ssa_nat_8datas_150dias_id_voo.xlsx"
URL_BASE = "https://www.google.com/travel/flights?hl=pt-BR&gl=BR&curr=BRL"

# Datas de referência fixas da coleta oficial.
# Inclui 45 e 90 dias para aumentar a população e melhorar a distribuição
# da variável antecedência, mantendo 120 e 150 dias para a janela longa.
INTERVALOS_REFERENCIA = [7, 15, 30, 45, 60, 90, 120, 150]
TOTAL_DATAS = len(INTERVALOS_REFERENCIA)
SEED_DATAS = 42
SEED_AMOSTRA = 2026
TAMANHO_AMOSTRA = 32

DEBUG_DIR = Path("debug_google_flights")
DEBUG_DIR.mkdir(exist_ok=True)

# =========================
# FUNÇÕES DE APOIO
# =========================


def esperar(driver, timeout=20):
    return WebDriverWait(driver, timeout)


def elementos(driver, by, seletor):
    try:
        return driver.find_elements(by, seletor)
    except Exception:
        return []


def clicar_js(driver, elemento):
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", elemento)
    time.sleep(0.3)
    driver.execute_script("arguments[0].click();", elemento)


def salvar_debug(driver, prefixo):
    """Salva screenshot e HTML para você inspecionar quando algo falhar."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    prefixo_limpo = re.sub(r"[^a-zA-Z0-9_-]", "_", prefixo)

    png = DEBUG_DIR / f"{prefixo_limpo}_{ts}.png"
    html = DEBUG_DIR / f"{prefixo_limpo}_{ts}.html"

    try:
        driver.save_screenshot(str(png))
    except Exception:
        pass

    try:
        html.write_text(driver.page_source, encoding="utf-8")
    except Exception:
        pass

    print(f"   Debug salvo em: {DEBUG_DIR.resolve()}")


def parsear_preco(texto):
    """Extrai preço em BRL. Ex.: R$ 1.234 -> 1234."""
    m = re.search(r"R\$[\s\u00a0\u202f]*([0-9\.]+)", texto)
    if not m:
        return None
    return int(m.group(1).replace(".", ""))


def parsear_duracao_min(texto):
    """Extrai duração em minutos. Ex.: 4 h 35 min -> 275."""
    m = re.search(r"(\d+)\s*h(?:\s*(\d+)\s*min)?", texto, flags=re.I)
    if m:
        horas = int(m.group(1))
        minutos = int(m.group(2)) if m.group(2) else 0
        return horas * 60 + minutos

    m = re.search(r"(\d+)\s*min", texto, flags=re.I)
    if m:
        return int(m.group(1))

    return None


def parsear_escalas(texto):
    texto_lower = texto.lower()

    if any(x in texto_lower for x in ["sem escala", "sem escalas", "direto", "nonstop"]):
        return 0

    m = re.search(r"(\d+)\s*(escala|escalas|parada|paradas|stop|stops)", texto_lower)
    if m:
        return int(m.group(1))

    # Melhor deixar vazio do que inventar 1 escala.
    return None


def parsear_companhia(texto):
    companhias = [
        "LATAM",
        "GOL",
        "Gol",
        "Azul",
        "Voepass",
        "Avianca",
        "Copa",
    ]

    for cia in companhias:
        if cia.lower() in texto.lower():
            return "GOL" if cia.lower() == "gol" else cia

    return None


def parsear_codigos_voo(texto):
    """
    Tenta extrair códigos operacionais dos voos no texto do card.

    Exemplos esperados:
        G3 1710
        LA 3350
        AD 4021
        2Z 2291

    Observação: o Google Flights nem sempre mostra o número do voo no card
    resumido. Quando não aparecer, o script deixa o código como vazio e usa
    um identificador operacional alternativo baseado em companhia/ordem/data.
    """
    padrao = re.compile(
        r"\b(?:G3|LA|AD|2Z|JJ|TP|CM|AV|AR|UX|IB|KL|AF|AA|DL|UA)\s*-?\s*\d{2,5}\b",
        flags=re.I,
    )

    encontrados = []
    vistos = set()

    for match in padrao.finditer(texto):
        codigo = re.sub(r"\s+", " ", match.group(0).upper().replace("-", " ")).strip()
        # Normaliza casos como G31710 -> G3 1710, LA3350 -> LA 3350, 2Z2291 -> 2Z 2291.
        codigo = re.sub(r"^(G3|LA|AD|2Z|JJ|TP|CM|AV|AR|UX|IB|KL|AF|AA|DL|UA)\s*(\d+)$", r"\1 \2", codigo)
        if codigo not in vistos:
            vistos.add(codigo)
            encontrados.append(codigo)

    return encontrados


def montar_identificador_operacional(codigos_voo, data_voo, companhia=None, ordem_card=None):
    """
    Monta um identificador descritivo para auditoria.

    Mantemos o id_populacao como identificador único oficial da linha.
    Este identificador operacional serve para reconhecer o voo observado.
    """
    data_txt = data_voo.strftime("%d/%m/%Y") if hasattr(data_voo, "strftime") else str(data_voo)

    if codigos_voo:
        return f"{' + '.join(codigos_voo)} - {data_txt}"

    partes = []
    if companhia:
        partes.append(companhia.upper())
    else:
        partes.append("VOO")

    if ordem_card is not None:
        partes.append(f"CARD-{ordem_card:03d}")

    return f"{' '.join(partes)} - {data_txt}"


def pedir_data_busca():
    while True:
        entrada = input("\nDigite a data de busca fixa (DD/MM/AAAA): ").strip()
        try:
            dia, mes, ano = entrada.split("/")
            data = date(int(ano), int(mes), int(dia))
            print(f"✓ Data de busca: {data.strftime('%d/%m/%Y')}")
            return data
        except Exception:
            print("  Formato inválido. Use DD/MM/AAAA")


def calcular_datas_viagem(data_busca):
    """
    Gera somente as datas fixas de referência.

    Ajuste feito para a coleta oficial: 8 datas fixas de antecedência
    (7, 15, 30, 45, 60, 90, 120 e 150 dias), aumentando a chance de
    atingir a população mínima de 620 unidades observacionais.
    """
    datas = [data_busca + timedelta(days=dias) for dias in sorted(INTERVALOS_REFERENCIA)]

    print(f"\n✓ {len(datas)} datas de viagem geradas:")
    for dt in datas:
        print(f"   {dt.strftime('%d/%m/%Y')} ({(dt - data_busca).days} dias)")

    populacao_estimada = len(datas) * 100
    print(f"\nEstimativa aproximada de população: {len(datas)} datas × ~100 voos ≈ {populacao_estimada} voos")

    return datas


def configurar_driver():
    options = Options()
    options.add_argument("--start-maximized")
    options.add_argument("--lang=pt-BR")

    # Não uso opções para mascarar automação. Se aparecer CAPTCHA/bloqueio,
    # pare a coleta e registre que a plataforma exigiu validação manual.
    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=options,
    )
    return driver


def aceitar_cookies(driver):
    textos = ["Aceitar tudo", "Aceitar", "Accept all", "I agree"]
    for texto in textos:
        try:
            botoes = driver.find_elements(By.XPATH, f'//button[contains(., "{texto}")]')
            if botoes:
                clicar_js(driver, botoes[0])
                time.sleep(1)
                return True
        except Exception:
            pass
    return False


def aguardar_resultados(driver, timeout=30):
    """Aguarda até aparecer algum card ou pelo menos algum preço em R$."""
    inicio = time.time()
    while time.time() - inicio < timeout:
        if coletar_elementos_candidatos(driver):
            return True
        time.sleep(0.7)
    return False


def abrir_calendario(driver):
    """Tenta abrir o calendário usando seletores e textos alternativos."""
    tentativas_css = [
        'input[aria-label="Partida"]',
        'input[aria-label="Data de partida"]',
        'input[aria-label*="partida" i]',
        'input[aria-label*="departure" i]',
        'input.TP4Lpb',
    ]

    for css in tentativas_css:
        try:
            el = esperar(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, css)))
            clicar_js(driver, el)
            time.sleep(1)
            return True
        except Exception:
            pass

    # Fallback por texto visível.
    tentativas_xpath = [
        '//*[contains(text(), "Partida")]',
        '//*[contains(text(), "Data de partida")]',
        '//*[contains(text(), "Departure")]',
    ]

    for xp in tentativas_xpath:
        try:
            els = driver.find_elements(By.XPATH, xp)
            if els:
                clicar_js(driver, els[0])
                time.sleep(1)
                return True
        except Exception:
            pass

    return False


def selecionar_data_calendario(driver, data_viagem):
    """Tenta selecionar a data no calendário. Retorna False se a UI mudou."""
    data_iso = data_viagem.strftime("%Y-%m-%d")
    dia_sem_zero = str(data_viagem.day)

    for _ in range(18):
        # 1) Melhor caso: elemento com data-iso.
        try:
            celulas = driver.find_elements(By.CSS_SELECTOR, f'[data-iso="{data_iso}"]')
            if celulas:
                clicar_js(driver, celulas[0])
                time.sleep(0.8)
                return True
        except Exception:
            pass

        # 2) Fallback por aria-label contendo o dia/mês/ano.
        try:
            ano = data_viagem.year
            candidatos = driver.find_elements(
                By.XPATH,
                f'//*[contains(@aria-label, "{ano}") and (contains(@aria-label, "{dia_sem_zero}") or text()="{dia_sem_zero}")]'
            )
            if candidatos:
                clicar_js(driver, candidatos[0])
                time.sleep(0.8)
                return True
        except Exception:
            pass

        # 3) Avança calendário.
        avancar_xpaths = [
            '//button[@aria-label="Avançar"]',
            '//button[@aria-label="Próximo"]',
            '//button[@aria-label="Next"]',
            '//button[contains(@aria-label, "próximo")]',
            '//button[contains(@aria-label, "Next")]',
        ]

        clicou_avancar = False
        for xp in avancar_xpaths:
            try:
                btns = driver.find_elements(By.XPATH, xp)
                if btns:
                    clicar_js(driver, btns[0])
                    time.sleep(0.8)
                    clicou_avancar = True
                    break
            except Exception:
                pass

        if not clicou_avancar:
            return False

    return False


def confirmar_data(driver):
    textos = ["Concluído", "Done", "OK"]
    for texto in textos:
        try:
            botoes = driver.find_elements(By.XPATH, f'//button[contains(., "{texto}")]')
            if botoes:
                clicar_js(driver, botoes[-1])
                time.sleep(1)
                return True
        except Exception:
            pass
    return False


def clicar_pesquisar(driver):
    tentativas = [
        (By.CSS_SELECTOR, '[aria-label="Pesquisar"]'),
        (By.CSS_SELECTOR, '[aria-label="Search"]'),
        (By.XPATH, '//button[contains(., "Pesquisar")]'),
        (By.XPATH, '//button[contains(., "Search")]'),
    ]

    for by, seletor in tentativas:
        try:
            botoes = driver.find_elements(by, seletor)
            if botoes:
                clicar_js(driver, botoes[0])
                time.sleep(2)
                return True
        except Exception:
            pass

    return False


def clicar_mais_voos(driver):
    textos = [
        "Mais voos",
        "Mostrar mais voos",
        "Ver mais voos",
        "More flights",
        "Show more flights",
    ]

    for _ in range(8):
        clicou = False
        for texto in textos:
            try:
                botoes = driver.find_elements(By.XPATH, f'//button[contains(., "{texto}")]')
                botoes = [b for b in botoes if b.is_displayed()]
                if botoes:
                    clicar_js(driver, botoes[0])
                    time.sleep(2)
                    clicou = True
                    break
            except (ElementClickInterceptedException, StaleElementReferenceException):
                time.sleep(1)
            except Exception:
                pass
        if not clicou:
            break


def coletar_elementos_candidatos(driver):
    """
    Busca cards por seletores conhecidos. Se não achar, usa fallbacks mais gerais.
    A deduplicação por texto evita contar o mesmo card várias vezes.
    """
    seletores = [
        'li.pIav2d',
        'li[role="listitem"]',
        'div[role="listitem"]',
    ]

    candidatos = []
    for css in seletores:
        candidatos.extend(elementos(driver, By.CSS_SELECTOR, css))

    vistos = set()
    cards = []

    for el in candidatos:
        try:
            texto = el.text.strip()
        except StaleElementReferenceException:
            continue

        if not texto:
            continue

        # Um card útil precisa ter preço e duração.
        if "R$" not in texto:
            continue
        if not re.search(r"\d+\s*h|\d+\s*min", texto, flags=re.I):
            continue

        normalizado = re.sub(r"\s+", " ", texto)
        if normalizado in vistos:
            continue

        vistos.add(normalizado)
        cards.append(el)

    return cards


def coletar_cards(driver, data_busca, data_viagem):
    antecedencia = (data_viagem - data_busca).days
    voos = []

    if not aguardar_resultados(driver, timeout=35):
        print("⚠ Não encontrei cards com preço/duração na tela.")
        salvar_debug(driver, f"sem_resultados_{data_viagem.isoformat()}")
        return []

    time.sleep(1.5)
    clicar_mais_voos(driver)
    time.sleep(1)

    cards = coletar_elementos_candidatos(driver)
    qtd_total = len(cards)

    for ordem, card in enumerate(cards, start=1):
        try:
            texto = card.text.strip()
            texto_bruto = re.sub(r"\n+", " | ", texto)

            preco = parsear_preco(texto)
            duracao_min = parsear_duracao_min(texto)
            escalas = parsear_escalas(texto)
            companhia = parsear_companhia(texto)
            codigos_voo = parsear_codigos_voo(texto)
            codigo_voo = " + ".join(codigos_voo) if codigos_voo else None
            identificador_operacional = montar_identificador_operacional(
                codigos_voo=codigos_voo,
                data_voo=data_viagem,
                companhia=companhia,
                ordem_card=ordem,
            )

            if preco is None:
                continue

            # Filtro simples para remover ruídos de preço que não sejam passagem.
            if preco <= 50:
                continue

            voos.append({
                "data_busca": data_busca.strftime("%d/%m/%Y"),
                "hora_coleta": datetime.now().strftime("%H:%M:%S"),
                "data_voo": data_viagem.strftime("%d/%m/%Y"),
                "antecedencia": antecedencia,
                "preco": preco,
                "duracao_min": duracao_min,
                "escalas": escalas,
                "companhia": companhia,
                "codigo_voo": codigo_voo,
                "identificador_operacional": identificador_operacional,
                "qtd_voos_data": qtd_total,
                "ordem_card": ordem,
                "fonte": "Google Flights",
                "texto_bruto": texto_bruto,
            })

        except Exception as e:
            print(f"   ⚠ Erro ao ler um card: {e}")
            continue

    print(f"✓ {len(voos)} voos coletados ({qtd_total} cards candidatos)")
    return voos


def trocar_data_ou_pedir_manual(driver, data_viagem):
    """
    Tenta trocar a data automaticamente.
    Se falhar, pede ajuste manual, sem pular a data.
    """
    print(f"   Tentando trocar data automaticamente para {data_viagem.strftime('%d/%m/%Y')}...")

    try:
        if abrir_calendario(driver):
            if selecionar_data_calendario(driver, data_viagem):
                confirmar_data(driver)
                clicar_pesquisar(driver)
                if aguardar_resultados(driver, timeout=25):
                    return True
    except Exception as e:
        print(f"   ⚠ Falha na troca automática: {e}")

    salvar_debug(driver, f"falha_troca_data_{data_viagem.isoformat()}")

    print("   ⚠ A troca automática da data falhou.")
    print(f"   Ajuste MANUALMENTE no navegador para: {data_viagem.strftime('%d/%m/%Y')}")
    print("   Garanta que está em: somente ida | SSA -> NAT | 1 adulto | econômica | BRL")
    input("   Quando os resultados aparecerem, pressione ENTER aqui para coletar essa data...")
    return True


def salvar_excel(todos_voos, arquivo, datas_planejadas, erros):
    wb = Workbook()

    fill_header = PatternFill("solid", fgColor="2F5496")
    fill_alt = PatternFill("solid", fgColor="DCE6F1")
    font_header = Font(color="FFFFFF", bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Aba 1: população completa
    ws = wb.active
    ws.title = "Populacao"

    cabecalhos = [
        "id_populacao",
        "ID Operacional",
        "Código do Voo",
        "Data Busca",
        "Hora Coleta",
        "Data Voo",
        "Antecedência (dias)",
        "Preço (R$)",
        "Duração (min)",
        "Escalas",
        "Companhia",
        "Qtd Voos na Data",
        "Ordem Card",
        "Fonte",
        "Texto Bruto",
    ]

    for col, titulo in enumerate(cabecalhos, start=1):
        c = ws.cell(1, col, titulo)
        c.fill = fill_header
        c.font = font_header
        c.alignment = align_center

    for i, voo in enumerate(todos_voos, start=1):
        valores = [
            i,
            voo.get("identificador_operacional"),
            voo.get("codigo_voo"),
            voo["data_busca"],
            voo["hora_coleta"],
            voo["data_voo"],
            voo["antecedencia"],
            voo["preco"],
            voo["duracao_min"],
            voo["escalas"],
            voo["companhia"],
            voo["qtd_voos_data"],
            voo["ordem_card"],
            voo["fonte"],
            voo["texto_bruto"],
        ]

        for col, valor in enumerate(valores, start=1):
            c = ws.cell(i + 1, col, valor)
            c.fill = fill_alt if i % 2 == 0 else PatternFill()
            c.alignment = align_left if col == 15 else align_center

    larguras = [14, 32, 24, 14, 12, 14, 20, 12, 16, 10, 18, 18, 12, 18, 90]
    for col, largura in enumerate(larguras, start=1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = largura

    # Aba 2: resumo por data
    ws2 = wb.create_sheet("Resumo_por_Data")
    resumo_headers = ["Data Voo", "Antecedência (dias)", "Nº Voos", "Preço Mín (R$)", "Preço Máx (R$)", "Preço Médio (R$)"]

    for col, titulo in enumerate(resumo_headers, start=1):
        c = ws2.cell(1, col, titulo)
        c.fill = fill_header
        c.font = font_header
        c.alignment = align_center

    por_data = defaultdict(list)
    for voo in todos_voos:
        por_data[voo["data_voo"]].append(voo)

    for linha, (data_voo, voos_data) in enumerate(sorted(por_data.items()), start=2):
        precos = [v["preco"] for v in voos_data if v["preco"]]
        valores = [
            data_voo,
            voos_data[0]["antecedencia"],
            len(voos_data),
            min(precos) if precos else None,
            max(precos) if precos else None,
            round(sum(precos) / len(precos), 2) if precos else None,
        ]
        for col, valor in enumerate(valores, start=1):
            ws2.cell(linha, col, valor).alignment = align_center

    for col in range(1, 7):
        ws2.column_dimensions[ws2.cell(1, col).column_letter].width = 22

    # Aba 3: amostra sorteada
    ws3 = wb.create_sheet("Amostra_32")
    for col, titulo in enumerate(cabecalhos, start=1):
        c = ws3.cell(1, col, titulo)
        c.fill = fill_header
        c.font = font_header
        c.alignment = align_center

    if len(todos_voos) >= TAMANHO_AMOSTRA:
        rng = random.Random(SEED_AMOSTRA)
        ids_sorteados = sorted(rng.sample(range(1, len(todos_voos) + 1), TAMANHO_AMOSTRA))
        mapa = {i + 1: voo for i, voo in enumerate(todos_voos)}

        for linha, id_pop in enumerate(ids_sorteados, start=2):
            voo = mapa[id_pop]
            valores = [
                id_pop,
                voo.get("identificador_operacional"),
                voo.get("codigo_voo"),
                voo["data_busca"],
                voo["hora_coleta"],
                voo["data_voo"],
                voo["antecedencia"],
                voo["preco"],
                voo["duracao_min"],
                voo["escalas"],
                voo["companhia"],
                voo["qtd_voos_data"],
                voo["ordem_card"],
                voo["fonte"],
                voo["texto_bruto"],
            ]
            for col, valor in enumerate(valores, start=1):
                c = ws3.cell(linha, col, valor)
                c.alignment = align_left if col == 15 else align_center
    else:
        ws3.cell(2, 1, f"População menor que {TAMANHO_AMOSTRA}. Não foi possível sortear a amostra.")

    for col, largura in enumerate(larguras, start=1):
        ws3.column_dimensions[ws3.cell(1, col).column_letter].width = largura

    # Aba 4: metodologia/log
    ws4 = wb.create_sheet("Log_Coleta")
    log_linhas = [
        ["Origem", ORIGEM],
        ["Destino", DESTINO],
        ["URL base", URL_BASE],
        ["Seed datas", SEED_DATAS],
        ["Seed amostra", SEED_AMOSTRA],
        ["Total de datas planejadas", len(datas_planejadas)],
        ["Total de voos coletados", len(todos_voos)],
        ["Tamanho da amostra", TAMANHO_AMOSTRA],
        ["Gerado em", datetime.now().strftime("%d/%m/%Y %H:%M:%S")],
    ]

    for linha, valores in enumerate(log_linhas, start=1):
        ws4.cell(linha, 1, valores[0]).font = Font(bold=True)
        ws4.cell(linha, 2, valores[1])

    linha_inicio_datas = len(log_linhas) + 3
    ws4.cell(linha_inicio_datas, 1, "Datas planejadas").font = Font(bold=True)
    ws4.cell(linha_inicio_datas + 1, 1, "Data")
    ws4.cell(linha_inicio_datas + 1, 2, "Antecedência")

    data_busca_ref = None
    if todos_voos:
        data_busca_ref = datetime.strptime(todos_voos[0]["data_busca"], "%d/%m/%Y").date()

    for idx, dt in enumerate(datas_planejadas, start=linha_inicio_datas + 2):
        ws4.cell(idx, 1, dt.strftime("%d/%m/%Y"))
        ws4.cell(idx, 2, (dt - data_busca_ref).days if data_busca_ref else "")

    if erros:
        linha_erros = linha_inicio_datas + 3 + len(datas_planejadas)
        ws4.cell(linha_erros, 1, "Ocorrências").font = Font(bold=True)
        for idx, erro in enumerate(erros, start=linha_erros + 1):
            ws4.cell(idx, 1, erro)

    ws4.column_dimensions["A"].width = 35
    ws4.column_dimensions["B"].width = 80

    wb.save(arquivo)
    print(f"\n✓ Arquivo salvo: {arquivo}")


def main():
    print("=" * 72)
    print("  SCRAPER ASSISTIDO GOOGLE FLIGHTS — SSA -> NAT")
    print("  Coleta população + sorteio de amostra")
    print("=" * 72)

    data_busca = pedir_data_busca()
    datas_viagem = calcular_datas_viagem(data_busca)

    print("\nO navegador vai abrir no Google Flights.")
    print("Configure manualmente a PRIMEIRA busca:")
    print(f"  Origem: {ORIGEM}")
    print(f"  Destino: {DESTINO}")
    print("  Tipo: somente ida")
    print("  Passageiros: 1 adulto")
    print("  Classe: econômica")
    print("  Moeda: BRL")
    print(f"  Data inicial: {datas_viagem[0].strftime('%d/%m/%Y')}")

    input("\nPressione ENTER para abrir o navegador...")

    driver = configurar_driver()
    driver.get(URL_BASE)
    time.sleep(4)
    aceitar_cookies(driver)

    input("\nDepois que a primeira busca estiver mostrando os resultados, pressione ENTER aqui...")

    todos_voos = []
    erros = []

    try:
        for indice, data_viagem in enumerate(datas_viagem, start=1):
            antecedencia = (data_viagem - data_busca).days
            print(f"\n[{indice}/{len(datas_viagem)}] Data {data_viagem.strftime('%d/%m/%Y')} | {antecedencia} dias")

            if indice > 1:
                trocar_data_ou_pedir_manual(driver, data_viagem)

            voos = coletar_cards(driver, data_busca, data_viagem)

            if not voos:
                msg = f"Nenhum voo coletado para {data_viagem.strftime('%d/%m/%Y')}"
                erros.append(msg)
                print(f"   ⚠ {msg}")
            else:
                todos_voos.extend(voos)

            # Salva parcial a cada data para evitar perder tudo se o navegador travar.
            if todos_voos:
                salvar_excel(todos_voos, ARQUIVO_SAIDA, datas_viagem, erros)

            if indice < len(datas_viagem):
                pausa = random.randint(4, 8)
                print(f"   Pausa de {pausa}s antes da próxima data...")
                time.sleep(pausa)

    finally:
        driver.quit()

    print("\n" + "=" * 72)
    print(f"Total de voos na população: {len(todos_voos)}")

    if todos_voos:
        salvar_excel(todos_voos, ARQUIVO_SAIDA, datas_viagem, erros)
        print("\nPronto. Verifique as abas:")
        print("  1. Populacao")
        print("  2. Resumo_por_Data")
        print("  3. Amostra_32")
        print("  4. Log_Coleta")
    else:
        print("\n⚠ Nenhum voo foi coletado. Veja a pasta debug_google_flights.")


if __name__ == "__main__":
    main()
